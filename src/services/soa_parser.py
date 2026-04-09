"""
Service de parsing des Schedule of Assessments (SoA) depuis fichiers Excel.

Supporte l'extraction automatique des visites, jours cibles et fenêtres
depuis un fichier Excel contenant un tableau SoA.
"""

import re
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from io import BytesIO


@dataclass
class VisitConfig:
    """Configuration d'une visite extraite du SoA."""
    visit_name: str
    target_day: int
    window_before: int
    window_after: int
    procedures: List[str] = None

    def __post_init__(self):
        if self.procedures is None:
            self.procedures = []

    def to_dict(self) -> Dict:
        return {
            "visit_name": self.visit_name,
            "target_day": self.target_day,
            "window_before": self.window_before,
            "window_after": self.window_after,
            "procedures": self.procedures,
        }


class SoaParserService:
    """
    Service de parsing des Schedule of Assessments depuis Excel.

    Détecte automatiquement :
    - La feuille contenant le SoA
    - Les visites (V1, V2, Screening, Baseline, EOT, etc.)
    - Les jours cibles (D0, D7, Day 14, J28, etc.)
    - Les fenêtres (±3, -2/+5, +/-3 days, etc.)
    - Les procédures (lignes avec X ou checkmarks)
    """

    # Patterns de noms de visites
    VISIT_PATTERNS = [
        r"^V\d+",                    # V1, V2, V10
        r"^Screening",              # Screening
        r"^Baseline",               # Baseline
        r"^EOT",                    # End of Treatment
        r"^EOS",                    # End of Study
        r"^FU",                     # Follow-up
        r"^Follow[\s-]?up",         # Follow-up
        r"^Visit\s*\d+",            # Visit 1, Visit2
        r"^W\d+",                   # W1, W4 (weeks)
        r"^D\d+",                   # D1, D14 (days as visit names)
        r"^J\d+",                   # J0, J7 (français)
        r"^Visite\s*\d+",           # Visite 1 (français)
    ]

    # Patterns de jours
    DAY_PATTERNS = [
        r"D\s*(-?\d+)",             # D0, D7, D-28
        r"Day\s*(-?\d+)",           # Day 0, Day 14
        r"J\s*(-?\d+)",             # J0, J7 (français)
        r"Jour\s*(-?\d+)",          # Jour 0 (français)
    ]

    # Patterns de fenêtres
    WINDOW_PATTERNS = [
        r"[±]\s*(\d+)",                     # ±3, ± 3
        r"\+\s*/\s*-\s*(\d+)",              # +/-3, +/- 3
        r"-\s*/\s*\+\s*(\d+)",              # -/+3
        r"-\s*(\d+)\s*/\s*\+\s*(\d+)",      # -2/+5, -2 / +5
        r"-\s*(\d+)\s*(?:to|à)\s*\+\s*(\d+)",  # -2 to +5, -2 à +5
        r"\(\s*-\s*(\d+)\s*,\s*\+\s*(\d+)\s*\)",  # (-2, +5)
        r"\(\s*-\s*(\d+)\s*/\s*\+\s*(\d+)\s*\)",  # (-2/+5)
    ]

    def __init__(self):
        self.visits: List[VisitConfig] = []
        self.procedures: List[str] = []

    def parse_file(self, file_bytes: bytes) -> List[VisitConfig]:
        """
        Parse un fichier Excel et extrait les configurations de visites.

        Args:
            file_bytes: Contenu du fichier Excel en bytes

        Returns:
            Liste des configurations de visites extraites
        """
        self.visits = []
        self.procedures = []

        # Charger le workbook
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)

        # Trouver la feuille contenant le SoA
        soa_sheet = self._find_soa_sheet(wb)
        if not soa_sheet:
            raise ValueError("Aucun tableau SoA trouvé dans le fichier Excel")

        # Parser le SoA
        self._parse_soa_sheet(soa_sheet)

        return self.visits

    def _find_soa_sheet(self, wb) -> Optional[Worksheet]:
        """
        Trouve la feuille contenant le Schedule of Assessments.

        Recherche par nom de feuille puis par contenu.
        """
        # Noms de feuilles courants pour SoA
        soa_names = ["soa", "schedule", "assessments", "visits", "visites", "planning"]

        # Recherche par nom
        for sheet_name in wb.sheetnames:
            if any(name in sheet_name.lower() for name in soa_names):
                return wb[sheet_name]

        # Recherche par contenu (présence de visites)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if self._sheet_contains_soa(sheet):
                return sheet

        # Retourner la première feuille si rien trouvé
        if wb.sheetnames:
            return wb[wb.sheetnames[0]]

        return None

    def _sheet_contains_soa(self, sheet: Worksheet) -> bool:
        """Vérifie si une feuille contient un tableau SoA."""
        visit_count = 0
        for row in sheet.iter_rows(min_row=1, max_row=20, max_col=50, values_only=True):
            for cell in row:
                if cell and self._is_visit_header(str(cell)):
                    visit_count += 1
                    if visit_count >= 3:
                        return True
        return False

    def _is_visit_header(self, text: str) -> bool:
        """Vérifie si un texte correspond à un nom de visite."""
        text = text.strip()
        for pattern in self.VISIT_PATTERNS:
            if re.match(pattern, text, re.IGNORECASE):
                return True
        return False

    def _parse_soa_sheet(self, sheet: Worksheet) -> None:
        """Parse une feuille SoA pour extraire les visites."""
        # Détecter le format et dispatcher
        if self._is_format_week(sheet):
            self._parse_format_week(sheet)
            return

        # Trouver la ligne header avec les visites (Format 1 & 2)
        header_row, visit_cols = self._find_visit_header(sheet)
        if not header_row:
            raise ValueError("Impossible de trouver la ligne des visites")

        # Extraire les visites et leurs jours/fenêtres
        for col_idx, visit_info in visit_cols.items():
            visit_name = visit_info["name"]

            # Chercher le jour et la fenêtre dans les lignes suivantes ou dans le header
            target_day, window_before, window_after = self._extract_day_window(
                sheet, header_row, col_idx, visit_info
            )

            visit = VisitConfig(
                visit_name=visit_name,
                target_day=target_day,
                window_before=window_before,
                window_after=window_after,
            )
            self.visits.append(visit)

        # Extraire les procédures (lignes avec X)
        self._extract_procedures(sheet, header_row, visit_cols)

    def _is_format_week(self, sheet: Worksheet) -> bool:
        """
        Détecte le Format Week : ligne avec 'Week' en colonne 2 suivi de valeurs numériques.

        Exemple TRANSTELLAR :
            Col1='Clinic Visit', Col2='Week', Col3=0, Col4=8, Col5=16 ...
        """
        for row_idx in range(1, 10):
            col2 = sheet.cell(row=row_idx, column=2).value
            if col2 and str(col2).strip().lower() == "week":
                # Vérifier que les colonnes suivantes contiennent des entiers
                numeric_count = 0
                for col_idx in range(3, min(sheet.max_column + 1, 15)):
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (int, float)):
                        numeric_count += 1
                if numeric_count >= 3:
                    return True
        return False

    def _parse_format_week(self, sheet: Worksheet) -> None:
        """
        Parse le Format Week (ex: TRANSTELLAR).

        Structure :
            Ligne 1 : 'Clinic Visit' | 'Week' | 0 | 8 | 16 | ... | 'EOT'
            Ligne 2 : None | ' Window (days)' | None | '±5' | '±5' | ...
            Lignes suivantes : nom procédure | None | 'X' | None | ...
        """
        # Trouver la ligne header (col 2 == 'week')
        header_row = None
        for row_idx in range(1, 10):
            col2 = sheet.cell(row=row_idx, column=2).value
            if col2 and str(col2).strip().lower() == "week":
                header_row = row_idx
                break

        if not header_row:
            raise ValueError("Format Week : ligne 'Week' introuvable")

        # Trouver la ligne des fenêtres (col 2 contient 'window')
        window_row = None
        for row_idx in range(header_row + 1, header_row + 5):
            col2 = sheet.cell(row=row_idx, column=2).value
            if col2 and "window" in str(col2).strip().lower():
                window_row = row_idx
                break

        # Construire les configs de visites
        visit_cols: Dict[int, None] = {}  # col_idx -> None (pour _extract_procedures)
        last_numeric_day = 0

        for col_idx in range(3, sheet.max_column + 1):
            cell_val = sheet.cell(row=header_row, column=col_idx).value
            if cell_val is None:
                continue

            # Calculer le jour cible
            if isinstance(cell_val, (int, float)):
                week = int(cell_val)
                visit_name = f"W{week}"
                target_day = week * 7
                last_numeric_day = max(last_numeric_day, target_day)
            elif str(cell_val).strip().upper() in ("EOT", "EOS", "FU", "FOLLOW-UP"):
                visit_name = str(cell_val).strip()
                target_day = None  # Calculé après
            else:
                continue

            # Extraire la fenêtre depuis la ligne Window
            window_before, window_after = 0, 0
            if window_row:
                win_val = sheet.cell(row=window_row, column=col_idx).value
                if win_val is not None:
                    if isinstance(win_val, (int, float)):
                        window_before = window_after = int(win_val)
                    else:
                        window = self._parse_window(str(win_val))
                        if window:
                            window_before, window_after = window
                        else:
                            # Fallback : extraire uniquement le nombre
                            num_match = re.search(r"(\d+)", str(win_val))
                            if num_match:
                                val = int(num_match.group(1))
                                window_before = window_after = val

            visit_cols[col_idx] = None
            self.visits.append(VisitConfig(
                visit_name=visit_name,
                target_day=target_day,  # type: ignore[arg-type]
                window_before=window_before,
                window_after=window_after,
            ))

        # Résoudre les jours cibles des visites spéciales (EOT etc.)
        for visit in self.visits:
            if visit.target_day is None:
                visit.target_day = last_numeric_day

        # Extraire les procédures
        self._extract_procedures(sheet, header_row, visit_cols)

    def _find_visit_header(self, sheet: Worksheet) -> Tuple[int, Dict]:
        """
        Trouve la ligne contenant les headers de visites.

        Returns:
            Tuple (numéro de ligne, dict {col_idx: {"name": visit_name, "raw": raw_text}})
        """
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
            visit_cols = {}
            for col_idx, cell in enumerate(row, start=1):
                if cell and self._is_visit_header(str(cell)):
                    raw_text = str(cell).strip()
                    visit_name = self._extract_visit_name(raw_text)
                    visit_cols[col_idx] = {
                        "name": visit_name,
                        "raw": raw_text,
                    }

            # Si on a trouvé au moins 3 visites, c'est probablement le header
            if len(visit_cols) >= 3:
                return row_idx, visit_cols

        return None, {}

    def _extract_visit_name(self, raw_text: str) -> str:
        """
        Extrait le nom propre de la visite depuis le texte brut.

        Exemples:
            "V1 D0" -> "V1"
            "V2 D7±2" -> "V2"
            "Screening (Day -28)" -> "Screening"
            "Baseline D0" -> "Baseline"
        """
        text = raw_text.strip()

        # Pattern Format 2: nom suivi de D/Day/J et un nombre
        match = re.match(r"^([A-Za-z]+\s*\d*)\s*[DJ]", text, re.IGNORECASE)
        if match:
            return match.group(1).strip()

        match = re.match(r"^([A-Za-z]+\s*\d*)\s*\(", text, re.IGNORECASE)
        if match:
            return match.group(1).strip()

        # Pattern simple: juste le nom
        for pattern in self.VISIT_PATTERNS:
            match = re.match(pattern, text, re.IGNORECASE)
            if match:
                return match.group(0).strip()

        return text

    def _extract_day_window(self, sheet: Worksheet, header_row: int, col_idx: int,
                            visit_info: Dict) -> Tuple[int, int, int]:
        """
        Extrait le jour cible et les fenêtres pour une visite.

        Cherche d'abord dans le texte du header (Format 2: "V1 D0"),
        puis dans les lignes Day/Window en dessous.

        Returns:
            Tuple (target_day, window_before, window_after)
        """
        raw_text = visit_info["raw"]
        target_day = 0
        window_before = 0
        window_after = 0

        # 1. Essayer d'extraire depuis le header (Format 2)
        day_match = None
        for pattern in self.DAY_PATTERNS:
            day_match = re.search(pattern, raw_text, re.IGNORECASE)
            if day_match:
                target_day = int(day_match.group(1))
                break

        window = self._parse_window(raw_text)
        if window:
            window_before, window_after = window

        # 2. Chercher dans les lignes suivantes (Format 1)
        for row_offset in range(1, 5):
            row_idx = header_row + row_offset
            if row_idx > sheet.max_row:
                break

            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if not cell_value:
                continue

            cell_text = str(cell_value).strip()

            # Chercher le jour si pas encore trouvé
            if target_day == 0:
                for pattern in self.DAY_PATTERNS:
                    day_match = re.search(pattern, cell_text, re.IGNORECASE)
                    if day_match:
                        target_day = int(day_match.group(1))
                        break

            # Chercher la fenêtre si pas encore trouvée
            if window_before == 0 and window_after == 0:
                window = self._parse_window(cell_text)
                if window:
                    window_before, window_after = window

            # Vérifier aussi le label de la ligne (première colonne)
            label = sheet.cell(row=row_idx, column=1).value
            if label:
                label_text = str(label).strip().lower()
                if "day" in label_text or "jour" in label_text:
                    # C'est une ligne Day
                    if target_day == 0:
                        for pattern in self.DAY_PATTERNS:
                            day_match = re.search(pattern, cell_text, re.IGNORECASE)
                            if day_match:
                                target_day = int(day_match.group(1))
                                break
                        # Peut-être juste un nombre
                        if target_day == 0:
                            num_match = re.match(r"^(-?\d+)$", cell_text)
                            if num_match:
                                target_day = int(num_match.group(1))

                elif "window" in label_text or "fenêtre" in label_text or "fenetre" in label_text:
                    # C'est une ligne Window
                    window = self._parse_window(cell_text)
                    if window:
                        window_before, window_after = window

        return target_day, window_before, window_after

    def _parse_window(self, text: str) -> Optional[Tuple[int, int]]:
        """
        Parse une fenêtre depuis un texte.

        Exemples:
            "±3" -> (3, 3)
            "+/-3" -> (3, 3)
            "-2/+5" -> (2, 5)
            "-2 to +5" -> (2, 5)

        Returns:
            Tuple (window_before, window_after) ou None si non trouvé
        """
        text = text.strip()

        # Pattern ±X ou +/-X
        match = re.search(r"[±]\s*(\d+)", text)
        if match:
            val = int(match.group(1))
            return (val, val)

        match = re.search(r"\+\s*/\s*-\s*(\d+)", text)
        if match:
            val = int(match.group(1))
            return (val, val)

        match = re.search(r"-\s*/\s*\+\s*(\d+)", text)
        if match:
            val = int(match.group(1))
            return (val, val)

        # Pattern asymétrique -X/+Y
        match = re.search(r"-\s*(\d+)\s*/\s*\+\s*(\d+)", text)
        if match:
            return (int(match.group(1)), int(match.group(2)))

        match = re.search(r"-\s*(\d+)\s*(?:to|à)\s*\+\s*(\d+)", text, re.IGNORECASE)
        if match:
            return (int(match.group(1)), int(match.group(2)))

        match = re.search(r"\(\s*-\s*(\d+)\s*[,/]\s*\+\s*(\d+)\s*\)", text)
        if match:
            return (int(match.group(1)), int(match.group(2)))

        return None

    def _extract_procedures(self, sheet: Worksheet, header_row: int, visit_cols: Dict) -> None:
        """Extrait les procédures du SoA (lignes avec X ou checkmarks)."""
        self.procedures = []

        for row_idx in range(header_row + 1, min(header_row + 100, sheet.max_row + 1)):
            # Récupérer le nom de la procédure (première colonne non vide)
            procedure_name = None
            for col_idx in range(1, min(5, sheet.max_column + 1)):
                cell = sheet.cell(row=row_idx, column=col_idx).value
                if cell and str(cell).strip():
                    text = str(cell).strip().lower()
                    # Ignorer les lignes Day/Window/Header
                    if any(kw in text for kw in ["day", "jour", "window", "fenêtre", "visite", "visit"]):
                        break
                    procedure_name = str(cell).strip()
                    break

            if not procedure_name:
                continue

            # Vérifier si au moins une visite a un X ou checkmark
            has_mark = False
            for col_idx in visit_cols.keys():
                cell = sheet.cell(row=row_idx, column=col_idx).value
                if cell:
                    cell_text = str(cell).strip().upper()
                    if cell_text in ["X", "x", "✓", "✔", "√", "YES", "OUI", "1"]:
                        has_mark = True
                        break

            if has_mark:
                self.procedures.append(procedure_name)

    def get_procedures(self) -> List[str]:
        """Retourne la liste des procédures extraites."""
        return self.procedures
