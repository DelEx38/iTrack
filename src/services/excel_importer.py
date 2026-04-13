"""
Service d'import de données Excel (format iTrack) vers SQLite.

Sheets reconnus :
- "Settings"                     → visit_config (configurations de visites)
- "Suivi Patients"               → patients + visits
- "Événements Indésirables"      → adverse_events
"""

from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple
import sqlite3


@dataclass
class ImportPreview:
    """Résumé des données détectées avant import."""
    patients: int = 0
    visits: int = 0
    adverse_events: int = 0
    visit_configs: int = 0
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def has_data(self) -> bool:
        return bool(self.patients or self.visits or self.adverse_events or self.visit_configs)


@dataclass
class ImportResult:
    """Résultat de l'import."""
    patients_created: int = 0
    patients_skipped: int = 0
    visits_created: int = 0
    visits_skipped: int = 0
    ae_created: int = 0
    ae_skipped: int = 0
    visit_configs_updated: int = 0
    errors: List[str] = field(default_factory=list)

    def summary(self) -> str:
        parts = []
        if self.patients_created:
            parts.append(f"{self.patients_created} patients")
        if self.visits_created:
            parts.append(f"{self.visits_created} visites")
        if self.ae_created:
            parts.append(f"{self.ae_created} EI")
        if self.visit_configs_updated:
            parts.append(f"{self.visit_configs_updated} configs visites")
        if not parts:
            return "Aucune donnée importée"
        skipped = self.patients_skipped + self.visits_skipped + self.ae_skipped
        s = "Importé : " + ", ".join(parts)
        if skipped:
            s += f" ({skipped} ignorés, déjà présents)"
        return s


class ExcelImporter:
    """
    Service d'import depuis un fichier Excel iTrack vers SQLite.

    Usage::

        importer = ExcelImporter()
        preview  = importer.preview(file_bytes)
        result   = importer.import_data(file_bytes, study_id, conn,
                                        on_conflict="skip")
    """

    # Mapping statuts patient Excel → SQLite
    _STATUS_MAP = {
        "Screening": "Screening",
        "Inclus": "Active",
        "Actif": "Active",
        "Terminé": "Completed",
        "Sorti": "Withdrawn",
    }

    # Mapping sévérité AE
    _SEVERITY_MAP = {
        "Léger": "Mild",
        "Modéré": "Moderate",
        "Sévère": "Severe",
    }

    # Mapping outcome AE
    _OUTCOME_MAP = {
        "En cours": "Ongoing",
        "Résolu": "Recovered",
        "Résolu avec séquelles": "Recovered with sequelae",
        "Décès": "Fatal",
    }

    # Mapping causalité AE
    _CAUSALITY_MAP = {
        "Non lié": "Not related",
        "Peu probable": "Unlikely",
        "Possible": "Possible",
        "Probable": "Probable",
        "Certain": "Certain",
    }

    # ─── API publique ─────────────────────────────────────────────────────────

    def preview(self, file_bytes: bytes) -> ImportPreview:
        """Analyse le fichier et retourne les stats sans rien importer."""
        try:
            from openpyxl import load_workbook
            from io import BytesIO
        except ImportError:
            p = ImportPreview()
            p.errors.append("openpyxl non disponible — pip install openpyxl")
            return p

        result = ImportPreview()
        try:
            wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
            names_lower = [s.lower() for s in wb.sheetnames]

            if "settings" in names_lower:
                ws = wb[wb.sheetnames[names_lower.index("settings")]]
                result.visit_configs = self._count_visit_configs(ws)

            patients_ws = self._find_sheet(wb, names_lower, ["suivi patients", "patients"])
            if patients_ws:
                result.patients, result.visits = self._count_patients_visits(patients_ws)

            ae_ws = self._find_sheet(
                wb, names_lower,
                ["événements indésirables", "evenements indesirables", "ae"]
            )
            if ae_ws:
                result.adverse_events = self._count_ae(ae_ws)

            if not result.has_data:
                result.warnings.append(
                    "Aucun onglet reconnu (attendu : 'Settings', 'Suivi Patients', "
                    "'Événements Indésirables')"
                )
            wb.close()
        except Exception as ex:
            result.errors.append(f"Erreur lecture fichier : {ex}")

        return result

    def import_data(
        self,
        file_bytes: bytes,
        study_id: int,
        conn: sqlite3.Connection,
        import_visit_configs: bool = True,
        import_patients: bool = True,
        import_ae: bool = True,
        on_conflict: str = "skip",
    ) -> ImportResult:
        """
        Importe les données dans la base SQLite.

        Args:
            file_bytes:           Contenu du fichier .xlsx
            study_id:             ID de l'étude cible dans la DB
            conn:                 Connexion SQLite active
            import_visit_configs: Importer/màj les configs de visites
            import_patients:      Importer patients et visites réalisées
            import_ae:            Importer les événements indésirables
            on_conflict:          "skip" → ignorer les doublons
                                  "update" → mettre à jour les enregistrements existants
        """
        from openpyxl import load_workbook
        from io import BytesIO

        result = ImportResult()
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        names_lower = [s.lower() for s in wb.sheetnames]

        try:
            if import_visit_configs and "settings" in names_lower:
                ws = wb[wb.sheetnames[names_lower.index("settings")]]
                self._import_visit_configs(ws, study_id, conn, result)

            if import_patients:
                patients_ws = self._find_sheet(wb, names_lower, ["suivi patients", "patients"])
                if patients_ws:
                    visit_configs = self._load_visit_config_map(conn, study_id)
                    self._import_patients(patients_ws, study_id, conn, visit_configs, on_conflict, result)

            if import_ae:
                ae_ws = self._find_sheet(
                    wb, names_lower,
                    ["événements indésirables", "evenements indesirables", "ae"]
                )
                if ae_ws:
                    patient_map = self._load_patient_map(conn, study_id)
                    self._import_ae(ae_ws, conn, patient_map, on_conflict, result)

        finally:
            wb.close()

        return result

    # ─── Helpers génériques ───────────────────────────────────────────────────

    def _find_sheet(self, wb, names_lower: List[str], candidates: List[str]):
        for candidate in candidates:
            if candidate in names_lower:
                return wb[wb.sheetnames[names_lower.index(candidate)]]
        return None

    def _parse_date(self, value) -> Optional[date]:
        """Convertit une valeur de cellule en date Python."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            # Numéro de série Excel
            try:
                from datetime import timedelta
                return (datetime(1899, 12, 30) + timedelta(days=int(value))).date()
            except Exception:
                return None
        if isinstance(value, str):
            s = value.strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%b-%y", "%d-%b-%Y", "%d/%m/%y"):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
        return None

    def _str(self, value) -> str:
        return str(value).strip() if value is not None else ""

    # ─── Comptage pour preview ────────────────────────────────────────────────

    def _count_visit_configs(self, ws) -> int:
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).startswith("V") and row[1] is not None:
                count += 1
        return count

    def _count_patients_visits(self, ws) -> Tuple[int, int]:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return 0, 0
        header = rows[0]
        visit_cols = [
            i for i, h in enumerate(header)
            if h and str(h).startswith("V") and str(h)[1:].isdigit()
        ]
        patients, visits = 0, 0
        for row in rows[1:]:
            if not row[0]:
                continue
            patients += 1
            for ci in visit_cols:
                if ci < len(row) and self._parse_date(row[ci]):
                    visits += 1
        return patients, visits

    def _count_ae(self, ws) -> int:
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            # N° Patient (col A) + Description (col E) + Date survenue (col C)
            if row[0] and row[4] and row[2]:
                count += 1
        return count

    # ─── Chargement depuis la DB ──────────────────────────────────────────────

    def _load_visit_config_map(self, conn: sqlite3.Connection, study_id: int) -> Dict[str, Dict]:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM visit_config WHERE study_id = ? ORDER BY visit_order",
            (study_id,)
        )
        return {row["visit_name"]: dict(row) for row in cursor.fetchall()}

    def _load_patient_map(self, conn: sqlite3.Connection, study_id: int) -> Dict[str, int]:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, patient_number FROM patients WHERE study_id = ?",
            (study_id,)
        )
        return {row["patient_number"]: row["id"] for row in cursor.fetchall()}

    # ─── Import Settings ──────────────────────────────────────────────────────

    def _import_visit_configs(
        self, ws, study_id: int, conn: sqlite3.Connection, result: ImportResult
    ):
        cursor = conn.cursor()
        existing = self._load_visit_config_map(conn, study_id)

        for row in ws.iter_rows(min_row=2, values_only=True):
            visit_name = self._str(row[0]) if row[0] else ""
            if not visit_name.startswith("V"):
                continue
            try:
                target_day = int(row[1]) if row[1] is not None else 0
                window_before = int(row[2]) if row[2] is not None else 0
                window_after = int(row[3]) if row[3] is not None else 0
            except (ValueError, TypeError):
                continue

            if visit_name in existing:
                cursor.execute(
                    "UPDATE visit_config SET target_day=?, window_before=?, window_after=? WHERE id=?",
                    (target_day, window_before, window_after, existing[visit_name]["id"])
                )
            else:
                cursor.execute(
                    "SELECT COALESCE(MAX(visit_order), 0) + 1 FROM visit_config WHERE study_id = ?",
                    (study_id,)
                )
                visit_order = cursor.fetchone()[0]
                cursor.execute(
                    """INSERT INTO visit_config
                       (study_id, visit_name, visit_order, target_day, window_before, window_after)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    (study_id, visit_name, visit_order, target_day, window_before, window_after)
                )
            result.visit_configs_updated += 1

        conn.commit()

    # ─── Import Patients + Visites ────────────────────────────────────────────

    def _import_patients(
        self, ws, study_id: int, conn: sqlite3.Connection,
        visit_configs: Dict, on_conflict: str, result: ImportResult
    ):
        # Détecter les colonnes de visites dans l'en-tête
        header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        visit_col_map: Dict[int, str] = {}
        for i, h in enumerate(header):
            if h and str(h).startswith("V") and str(h)[1:].isdigit():
                visit_col_map[i] = str(h)

        cursor = conn.cursor()

        for row in ws.iter_rows(min_row=2, values_only=True):
            patient_number = self._str(row[0]) if row[0] else ""
            if not patient_number:
                continue

            initials = self._str(row[1])
            birth_date = self._parse_date(row[2])
            consent_date = self._parse_date(row[3])
            randomization_arm = self._str(row[5])
            randomization_number = self._str(row[6])
            raw_status = self._str(row[7]) or "Screening"
            status = self._STATUS_MAP.get(raw_status, raw_status)
            exit_date = self._parse_date(row[8])
            exit_reason = self._str(row[9])

            # Rechercher le patient
            cursor.execute(
                "SELECT id FROM patients WHERE study_id = ? AND patient_number = ?",
                (study_id, patient_number)
            )
            existing = cursor.fetchone()

            if existing:
                patient_id = existing["id"]
                if on_conflict == "update":
                    cursor.execute(
                        """UPDATE patients
                           SET initials=?, birth_date=?, screening_date=?,
                               randomization_arm=?, randomization_number=?,
                               status=?, exit_date=?, exit_reason=?,
                               updated_at=CURRENT_TIMESTAMP
                           WHERE id=?""",
                        (initials, birth_date, consent_date, randomization_arm,
                         randomization_number, status, exit_date, exit_reason, patient_id)
                    )
                    result.patients_created += 1
                else:
                    result.patients_skipped += 1
            else:
                cursor.execute(
                    """INSERT INTO patients
                       (study_id, patient_number, initials, birth_date, screening_date,
                        randomization_arm, randomization_number, status, exit_date, exit_reason)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (study_id, patient_number, initials, birth_date, consent_date,
                     randomization_arm, randomization_number, status, exit_date, exit_reason)
                )
                patient_id = cursor.lastrowid
                result.patients_created += 1

            conn.commit()

            # Visites
            for col_idx, visit_name in visit_col_map.items():
                if col_idx >= len(row):
                    continue
                visit_date = self._parse_date(row[col_idx])
                if not visit_date:
                    continue

                vc = visit_configs.get(visit_name)
                if not vc:
                    result.visits_skipped += 1
                    continue

                vc_id = vc["id"]
                cursor.execute(
                    "SELECT id FROM visits WHERE patient_id = ? AND visit_config_id = ?",
                    (patient_id, vc_id)
                )
                existing_visit = cursor.fetchone()

                if existing_visit:
                    if on_conflict == "update":
                        cursor.execute(
                            "UPDATE visits SET visit_date=?, status='Completed' WHERE id=?",
                            (visit_date, existing_visit["id"])
                        )
                        result.visits_created += 1
                    else:
                        result.visits_skipped += 1
                else:
                    cursor.execute(
                        """INSERT INTO visits (patient_id, visit_config_id, visit_date, status)
                           VALUES (?, ?, ?, 'Completed')""",
                        (patient_id, vc_id, visit_date)
                    )
                    result.visits_created += 1

                conn.commit()

    # ─── Import AEs ───────────────────────────────────────────────────────────

    def _import_ae(
        self, ws, conn: sqlite3.Connection,
        patient_map: Dict[str, int], on_conflict: str, result: ImportResult
    ):
        """
        Colonnes attendues (onglet "Événements Indésirables") :
        A  N° Patient          B  N° EI
        C  Date survenue       D  Date fin
        E  Description         F  Gravité
        G  Sévérité (grade)    H  Lien traitement (causalité)
        I  Action prise        J  Évolution (outcome)
        K  EIG (Oui/Non)       L  Date déclaration
        M  Commentaires
        """
        cursor = conn.cursor()

        for row in ws.iter_rows(min_row=2, values_only=True):
            patient_number = self._str(row[0]) if row[0] else ""
            description = self._str(row[4]) if row[4] else ""
            if not patient_number or not description:
                continue

            start_date = self._parse_date(row[2])
            if not start_date:
                continue

            patient_id = patient_map.get(patient_number)
            if not patient_id:
                result.ae_skipped += 1
                continue

            end_date = self._parse_date(row[3])
            severity = self._SEVERITY_MAP.get(self._str(row[5]), self._str(row[5]) or "Mild")
            causality = self._CAUSALITY_MAP.get(self._str(row[7]), self._str(row[7]) or "Unknown")
            outcome = self._OUTCOME_MAP.get(self._str(row[9]), self._str(row[9]) or "Ongoing")
            is_serious = self._str(row[10]).lower() == "oui" if row[10] else False
            reporting_date = self._parse_date(row[11])
            notes = self._str(row[12]) or None

            # Doublon : même patient + même date + même description
            cursor.execute(
                """SELECT id FROM adverse_events
                   WHERE patient_id = ? AND start_date = ? AND description = ?""",
                (patient_id, start_date, description)
            )
            existing = cursor.fetchone()

            if existing:
                if on_conflict == "update":
                    cursor.execute(
                        """UPDATE adverse_events
                           SET end_date=?, severity=?, causality=?, outcome=?,
                               is_serious=?, reporting_date=?, notes=?
                           WHERE id=?""",
                        (end_date, severity, causality, outcome,
                         int(is_serious), reporting_date, notes, existing["id"])
                    )
                    result.ae_created += 1
                else:
                    result.ae_skipped += 1
            else:
                cursor.execute(
                    "SELECT COALESCE(MAX(ae_number), 0) + 1 FROM adverse_events WHERE patient_id = ?",
                    (patient_id,)
                )
                ae_number = cursor.fetchone()[0]
                cursor.execute(
                    """INSERT INTO adverse_events
                       (patient_id, ae_number, ae_term, start_date, end_date, description,
                        severity, causality, outcome, is_serious, reporting_date, notes)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (patient_id, ae_number, description, start_date, end_date, description,
                     severity, causality, outcome, int(is_serious), reporting_date, notes)
                )
                result.ae_created += 1

            conn.commit()
