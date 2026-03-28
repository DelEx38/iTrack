"""
Fonctions de mise en forme avancée pour Excel.
"""

from typing import Optional, Literal, List, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule,
    FormulaRule, CellIsRule
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


# =============================================================================
# Types de bordures
# =============================================================================

BorderStyle = Literal["thin", "medium", "thick", "double", "dotted", "dashed"]


def create_border(
    style: BorderStyle = "thin",
    color: str = "000000",
    left: bool = True,
    right: bool = True,
    top: bool = True,
    bottom: bool = True
) -> Border:
    """
    Crée un objet Border personnalisé.

    Args:
        style: Style de bordure
        color: Couleur (hex sans #)
        left, right, top, bottom: Côtés à appliquer

    Returns:
        Objet Border
    """
    side = Side(style=style, color=color)
    return Border(
        left=side if left else None,
        right=side if right else None,
        top=side if top else None,
        bottom=side if bottom else None
    )


def apply_border(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    style: BorderStyle = "thin",
    color: str = "000000",
    outline_only: bool = False
) -> None:
    """
    Applique des bordures à une plage de cellules.

    Args:
        ws: Feuille de calcul
        start_row, start_col: Coin supérieur gauche
        end_row, end_col: Coin inférieur droit
        style: Style de bordure
        color: Couleur
        outline_only: Si True, bordure uniquement autour (pas à l'intérieur)
    """
    side = Side(style=style, color=color)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)

            if outline_only:
                # Bordure uniquement sur les bords extérieurs
                border = Border(
                    left=side if col == start_col else None,
                    right=side if col == end_col else None,
                    top=side if row == start_row else None,
                    bottom=side if row == end_row else None
                )
            else:
                # Bordure sur toutes les cellules
                border = Border(left=side, right=side, top=side, bottom=side)

            cell.border = border


def apply_border_range(
    ws: Worksheet,
    range_str: str,
    style: BorderStyle = "thin",
    color: str = "000000"
) -> None:
    """
    Applique des bordures à une plage (format A1:C10).

    Args:
        ws: Feuille de calcul
        range_str: Plage au format "A1:C10"
        style: Style de bordure
        color: Couleur
    """
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    apply_border(ws, min_row, min_col, max_row, max_col, style, color)


# =============================================================================
# Fusion de cellules
# =============================================================================

def merge_cells(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    value: Optional[str] = None,
    center: bool = True,
    bold: bool = False,
    bg_color: Optional[str] = None
) -> None:
    """
    Fusionne des cellules et applique un style optionnel.

    Args:
        ws: Feuille de calcul
        start_row, start_col: Coin supérieur gauche
        end_row, end_col: Coin inférieur droit
        value: Valeur à écrire dans la cellule fusionnée
        center: Centrer le texte
        bold: Texte en gras
        bg_color: Couleur de fond (hex sans #)
    """
    # Fusionner
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=end_row,
        end_column=end_col
    )

    # Appliquer le style à la cellule principale
    cell = ws.cell(row=start_row, column=start_col)

    if value is not None:
        cell.value = value

    if center:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if bold:
        cell.font = Font(bold=True)

    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")


def merge_range(
    ws: Worksheet,
    range_str: str,
    value: Optional[str] = None,
    center: bool = True
) -> None:
    """
    Fusionne une plage (format A1:C3).

    Args:
        ws: Feuille de calcul
        range_str: Plage au format "A1:C3"
        value: Valeur à écrire
        center: Centrer le texte
    """
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    merge_cells(ws, min_row, min_col, max_row, max_col, value, center)


# =============================================================================
# Formats de nombres
# =============================================================================

NUMBER_FORMATS = {
    "entier": "#,##0",
    "decimal": "#,##0.00",
    "pourcentage": "0.00%",
    "euro": '#,##0.00 €',
    "dollar": '$#,##0.00',
    "date": "DD/MM/YYYY",
    "date_heure": "DD/MM/YYYY HH:MM",
    "heure": "HH:MM:SS",
    "telephone": "00 00 00 00 00",
}


def apply_number_format(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    format_name: str
) -> None:
    """
    Applique un format de nombre à une plage.

    Args:
        ws: Feuille de calcul
        start_row, start_col: Coin supérieur gauche
        end_row, end_col: Coin inférieur droit
        format_name: Nom du format (voir NUMBER_FORMATS) ou format Excel personnalisé
    """
    fmt = NUMBER_FORMATS.get(format_name, format_name)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).number_format = fmt


def format_column(
    ws: Worksheet,
    col: int,
    format_name: str,
    start_row: int = 2
) -> None:
    """
    Applique un format à toute une colonne.

    Args:
        ws: Feuille de calcul
        col: Numéro de colonne (1-indexed)
        format_name: Nom du format
        start_row: Ligne de départ (défaut: 2 pour ignorer l'en-tête)
    """
    fmt = NUMBER_FORMATS.get(format_name, format_name)

    for row in range(start_row, ws.max_row + 1):
        ws.cell(row=row, column=col).number_format = fmt


# =============================================================================
# Alignement
# =============================================================================

def apply_alignment(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    horizontal: Literal["left", "center", "right", "justify"] = "left",
    vertical: Literal["top", "center", "bottom"] = "center",
    wrap_text: bool = False,
    indent: int = 0
) -> None:
    """
    Applique un alignement à une plage de cellules.

    Args:
        ws: Feuille de calcul
        start_row, start_col: Coin supérieur gauche
        end_row, end_col: Coin inférieur droit
        horizontal: Alignement horizontal
        vertical: Alignement vertical
        wrap_text: Retour à la ligne automatique
        indent: Indentation (en caractères)
    """
    alignment = Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap_text,
        indent=indent
    )

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).alignment = alignment


# =============================================================================
# Mise en forme conditionnelle
# =============================================================================

def add_color_scale(
    ws: Worksheet,
    range_str: str,
    min_color: str = "F8696B",  # Rouge
    mid_color: Optional[str] = "FFEB84",  # Jaune
    max_color: str = "63BE7B"  # Vert
) -> None:
    """
    Ajoute une échelle de couleurs (dégradé selon les valeurs).

    Args:
        ws: Feuille de calcul
        range_str: Plage au format "B2:B20"
        min_color: Couleur pour la valeur minimale
        mid_color: Couleur pour la valeur médiane (None = 2 couleurs seulement)
        max_color: Couleur pour la valeur maximale
    """
    if mid_color:
        rule = ColorScaleRule(
            start_type="min", start_color=min_color,
            mid_type="percentile", mid_value=50, mid_color=mid_color,
            end_type="max", end_color=max_color
        )
    else:
        rule = ColorScaleRule(
            start_type="min", start_color=min_color,
            end_type="max", end_color=max_color
        )

    ws.conditional_formatting.add(range_str, rule)


def add_data_bars(
    ws: Worksheet,
    range_str: str,
    color: str = "638EC6"
) -> None:
    """
    Ajoute des barres de données (visualisation dans les cellules).

    Args:
        ws: Feuille de calcul
        range_str: Plage au format "C2:C20"
        color: Couleur des barres
    """
    rule = DataBarRule(
        start_type="min",
        end_type="max",
        color=color,
        showValue=True,
        minLength=None,
        maxLength=None
    )
    ws.conditional_formatting.add(range_str, rule)


def add_icon_set(
    ws: Worksheet,
    range_str: str,
    icon_style: Literal["3Arrows", "3TrafficLights", "3Symbols", "4Arrows", "5Arrows"] = "3Arrows"
) -> None:
    """
    Ajoute des icônes conditionnelles.

    Args:
        ws: Feuille de calcul
        range_str: Plage
        icon_style: Style d'icônes
    """
    rule = IconSetRule(
        icon_style=icon_style,
        type="percent",
        values=[0, 33, 67] if icon_style.startswith("3") else [0, 25, 50, 75]
    )
    ws.conditional_formatting.add(range_str, rule)


def add_highlight_rule(
    ws: Worksheet,
    range_str: str,
    operator: Literal["greaterThan", "lessThan", "equal", "between", "containsText"],
    value: Any,
    bg_color: str = "FFFF00",
    font_color: str = "000000",
    bold: bool = False
) -> None:
    """
    Ajoute une règle de mise en surbrillance conditionnelle.

    Args:
        ws: Feuille de calcul
        range_str: Plage
        operator: Type de comparaison
        value: Valeur de comparaison (ou [min, max] pour "between")
        bg_color: Couleur de fond
        font_color: Couleur du texte
        bold: Texte en gras
    """
    fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    font = Font(color=font_color, bold=bold)

    if operator == "between" and isinstance(value, (list, tuple)):
        rule = CellIsRule(
            operator=operator,
            formula=value,
            fill=fill,
            font=font
        )
    else:
        rule = CellIsRule(
            operator=operator,
            formula=[value],
            fill=fill,
            font=font
        )

    ws.conditional_formatting.add(range_str, rule)


# =============================================================================
# Styles prédéfinis
# =============================================================================

def apply_style_preset(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    preset: Literal["tableau", "header", "total", "attention", "success", "warning"]
) -> None:
    """
    Applique un style prédéfini à une plage.

    Args:
        ws: Feuille de calcul
        start_row, start_col: Coin supérieur gauche
        end_row, end_col: Coin inférieur droit
        preset: Nom du style prédéfini
    """
    presets = {
        "header": {
            "font": Font(bold=True, color="FFFFFF"),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        },
        "total": {
            "font": Font(bold=True),
            "fill": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
            "border": create_border("medium"),
        },
        "attention": {
            "font": Font(bold=True, color="9C0006"),
            "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        },
        "success": {
            "font": Font(color="006100"),
            "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        },
        "warning": {
            "font": Font(color="9C5700"),
            "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        },
        "tableau": {
            "border": create_border("thin"),
            "alignment": Alignment(vertical="center"),
        },
    }

    style = presets.get(preset, {})

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            if "font" in style:
                cell.font = style["font"]
            if "fill" in style:
                cell.fill = style["fill"]
            if "alignment" in style:
                cell.alignment = style["alignment"]
            if "border" in style:
                cell.border = style["border"]
