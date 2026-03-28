"""
Templates spécifiques pour la recherche clinique.
"""

from typing import Optional, List, Dict
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

from .generator import save_workbook


# =============================================================================
# Constantes de style
# =============================================================================

DATE_FORMAT_US = "[$-409]DD-MMM-YY"

# Couleurs (hex sans #)
COLOR_BLUE = "4472C4"
COLOR_GREEN = "70AD47"
COLOR_GREEN_LIGHT = "C6EFCE"
COLOR_GREEN_PALE = "E2EFDA"
COLOR_GREEN_DARK = "00B050"
COLOR_RED = "C00000"
COLOR_RED_LIGHT = "FFC7CE"
COLOR_PURPLE = "7030A0"
COLOR_PURPLE_LIGHT = "9966FF"
COLOR_ORANGE = "ED7D31"

# Styles réutilisables
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")

# Fills réutilisables
FILL_BLUE = PatternFill(start_color=COLOR_BLUE, end_color=COLOR_BLUE, fill_type="solid")
FILL_GREEN = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
FILL_GREEN_LIGHT = PatternFill(start_color=COLOR_GREEN_LIGHT, end_color=COLOR_GREEN_LIGHT, fill_type="solid")
FILL_GREEN_PALE = PatternFill(start_color=COLOR_GREEN_PALE, end_color=COLOR_GREEN_PALE, fill_type="solid")
FILL_GREEN_DARK = PatternFill(start_color=COLOR_GREEN_DARK, end_color=COLOR_GREEN_DARK, fill_type="solid")
FILL_RED = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid")
FILL_RED_LIGHT = PatternFill(start_color=COLOR_RED_LIGHT, end_color=COLOR_RED_LIGHT, fill_type="solid")
FILL_PURPLE = PatternFill(start_color=COLOR_PURPLE, end_color=COLOR_PURPLE, fill_type="solid")
FILL_PURPLE_LIGHT = PatternFill(start_color=COLOR_PURPLE_LIGHT, end_color=COLOR_PURPLE_LIGHT, fill_type="solid")
FILL_ORANGE = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid")

# Configuration des consentements
CONSENT_CONFIGS = [
    {"name": "ICF_Principal", "display": "ICF Principal", "start_col": 6,
     "versions": ["v1.0", "v2.0", "v3.0", "", "", ""]},
    {"name": "ICF_PK", "display": "ICF Sous-étude PK", "start_col": 9,
     "versions": ["v1.0", "v2.0", "", "", "", ""]},
    {"name": "ICF_Genetique", "display": "ICF Génétique", "start_col": 12,
     "versions": ["v1.0", "", "", "", "", ""]},
]


# =============================================================================
# Fonctions utilitaires
# =============================================================================

def _apply_header_style(cell, fill: PatternFill = FILL_BLUE) -> None:
    """Applique le style d'en-tête à une cellule."""
    cell.font = HEADER_FONT
    cell.fill = fill
    cell.alignment = HEADER_ALIGN


def _create_table(ws: Worksheet, name: str, ref: str, style: str = "TableStyleMedium2") -> None:
    """Crée un tableau formaté."""
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)


def _set_column_widths(ws: Worksheet, widths: Dict[str, int]) -> None:
    """Définit les largeurs de colonnes."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# =============================================================================
# Création des onglets
# =============================================================================

def _create_settings_sheet(ws: Worksheet, num_visits: int) -> None:
    """Crée l'onglet Settings avec fenêtres de visite et versions consentements."""
    # En-têtes visites
    headers = ["Visite", "Jour cible", "Fenêtre -", "Fenêtre +"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    # Données visites
    for i in range(1, num_visits + 1):
        row = i + 1
        ws.cell(row=row, column=1, value=f"V{i}")
        # V1 = référence (jour 0, pas de fenêtre)
        if i == 1:
            ws.cell(row=row, column=2, value=0)
            ws.cell(row=row, column=3, value=0)
            ws.cell(row=row, column=4, value=0)
        else:
            for col in [2, 3, 4]:
                ws.cell(row=row, column=col, value="")

    _set_column_widths(ws, {"A": 10, "B": 12, "C": 12, "D": 12})
    _create_table(ws, "StudySettings", f"A1:D{num_visits + 1}")

    # Tableaux des versions de consentements
    for config in CONSENT_CONFIGS:
        _create_consent_table(ws, config)


def _create_consent_table(ws: Worksheet, config: Dict) -> None:
    """Crée un tableau de versions de consentement."""
    start_col = config["start_col"]
    versions = config["versions"]
    display_name = config["display"]

    # En-têtes
    headers = [f"{display_name} - Version", "Date", "Libellé"]
    for col_offset, header in enumerate(headers):
        cell = ws.cell(row=1, column=start_col + col_offset, value=header)
        _apply_header_style(cell, FILL_PURPLE)

    # Données
    for i, version in enumerate(versions, start=2):
        col_version = start_col
        col_date = start_col + 1
        col_libelle = start_col + 2

        ws.cell(row=i, column=col_version, value=version)
        ws.cell(row=i, column=col_date, value="")
        ws.cell(row=i, column=col_date).number_format = DATE_FORMAT_US

        # Formule Libellé
        col_v = get_column_letter(col_version)
        col_d = get_column_letter(col_date)
        formula = f'=IF({col_v}{i}<>"","{display_name} "&{col_v}{i}&" ("&TEXT({col_d}{i},"[$-409]DD-MMM-YY")&")","")'
        ws.cell(row=i, column=col_libelle, value=formula)

    # Largeurs et tableau
    ws.column_dimensions[get_column_letter(start_col)].width = 12
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 14
    ws.column_dimensions[get_column_letter(start_col + 2)].width = 32

    table_ref = f"{get_column_letter(start_col)}1:{get_column_letter(start_col + 2)}{len(versions) + 1}"
    _create_table(ws, config["name"], table_ref, "TableStyleMedium5")


def _create_patients_sheet(ws: Worksheet, num_visits: int, num_patients: int) -> None:
    """Crée l'onglet Suivi Patients."""
    fixed_headers = [
        "N° Patient", "Initiales", "Date naissance", "Date consentement",
        "Critères I/E OK", "Bras", "N° Randomisation", "Statut",
        "Date sortie", "Raison sortie"
    ]
    num_fixed = len(fixed_headers)

    # En-têtes fixes
    for col, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    # En-têtes visites
    for i in range(1, num_visits + 1):
        cell = ws.cell(row=1, column=num_fixed + i, value=f"V{i}")
        _apply_header_style(cell, FILL_GREEN)

    # Largeurs colonnes fixes
    widths = {"A": 12, "B": 10, "C": 12, "D": 14, "E": 14, "F": 10, "G": 16, "H": 10, "I": 12, "J": 18}
    _set_column_widths(ws, widths)

    # Largeurs colonnes visites
    for i in range(1, num_visits + 1):
        ws.column_dimensions[get_column_letter(num_fixed + i)].width = 12

    # Lignes vides et formats de date
    date_cols = [3, 4, 9]  # Date naissance, consentement, sortie
    for row in range(2, num_patients + 2):
        ws.cell(row=row, column=1, value="")
        for col in date_cols:
            ws.cell(row=row, column=col).number_format = DATE_FORMAT_US
        for i in range(1, num_visits + 1):
            ws.cell(row=row, column=num_fixed + i).number_format = DATE_FORMAT_US

    # Validations
    _add_validation(ws, "list", '"Oui,Non"', f"E2:E{num_patients + 1}", "Choisir Oui ou Non")
    _add_validation(ws, "list", '"Screening,Inclus,Actif,Terminé,Sorti"', f"H2:H{num_patients + 1}")
    _add_validation(ws, "list",
        '"Fin prévue,Retrait consentement,Perdu de vue,Événement indésirable,Décès,Décision investigateur,Autre"',
        f"J2:J{num_patients + 1}")

    # Tableau
    total_cols = num_fixed + num_visits
    _create_table(ws, "SuiviPatients", f"A1:{get_column_letter(total_cols)}{num_patients + 1}", "TableStyleMedium9")

    # Mise en forme conditionnelle (V2 à Vn)
    _add_visit_conditional_formatting(ws, num_visits, num_patients, num_fixed)

    ws.freeze_panes = f"{get_column_letter(num_fixed + 1)}2"


def _add_validation(ws: Worksheet, val_type: str, formula: str, range_str: str, error: str = None) -> None:
    """Ajoute une validation de données."""
    dv = DataValidation(type=val_type, formula1=formula, allow_blank=True)
    if error:
        dv.error = error
    ws.add_data_validation(dv)
    dv.add(range_str)


def _add_visit_conditional_formatting(ws: Worksheet, num_visits: int, num_patients: int, num_fixed: int) -> None:
    """Ajoute la mise en forme conditionnelle pour les visites."""
    v1_col_letter = get_column_letter(num_fixed + 1)

    for i in range(2, num_visits + 1):
        visit_col = num_fixed + i
        visit_col_letter = get_column_letter(visit_col)
        range_str = f"{visit_col_letter}2:{visit_col_letter}{num_patients + 1}"
        settings_row = i + 1

        # Formule verte (dans la fenêtre)
        formula_green = (
            f'AND({visit_col_letter}2<>"",${v1_col_letter}2<>"",'
            f'{visit_col_letter}2>=${v1_col_letter}2+Settings!$B${settings_row}+Settings!$C${settings_row},'
            f'{visit_col_letter}2<=${v1_col_letter}2+Settings!$B${settings_row}+Settings!$D${settings_row})'
        )
        # Formule rouge (hors fenêtre)
        formula_red = (
            f'AND({visit_col_letter}2<>"",${v1_col_letter}2<>"",'
            f'OR({visit_col_letter}2<${v1_col_letter}2+Settings!$B${settings_row}+Settings!$C${settings_row},'
            f'{visit_col_letter}2>${v1_col_letter}2+Settings!$B${settings_row}+Settings!$D${settings_row}))'
        )

        ws.conditional_formatting.add(range_str, FormulaRule(formula=[formula_green], fill=FILL_GREEN_LIGHT))
        ws.conditional_formatting.add(range_str, FormulaRule(formula=[formula_red], fill=FILL_RED_LIGHT))


def _create_documents_sheet(ws: Worksheet, num_patients: int) -> None:
    """Crée l'onglet Documents (consentements)."""
    consent_types = ["ICF Principal", "ICF Sous-étude PK", "ICF Génétique"]
    consent_table_names = ["ICF_Principal", "ICF_PK", "ICF_Genetique"]

    # En-têtes
    headers = ["N° Patient"]
    for ct in consent_types:
        headers.extend([f"{ct} - Version", f"{ct} - Date"])
    headers.extend(["Note info remise", "Commentaires"])

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        if "Version" in header:
            _apply_header_style(cell, FILL_PURPLE)
        elif "Date" in header:
            _apply_header_style(cell, FILL_PURPLE_LIGHT)
        else:
            _apply_header_style(cell)

    # Largeurs
    ws.column_dimensions["A"].width = 12
    col_idx = 2
    for _ in consent_types:
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 14
        col_idx += 2
    ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.column_dimensions[get_column_letter(col_idx + 1)].width = 30

    # Lignes et formats de date
    for row in range(2, num_patients + 2):
        ws.cell(row=row, column=1, value="")
        col_idx = 3
        for _ in consent_types:
            ws.cell(row=row, column=col_idx).number_format = DATE_FORMAT_US
            col_idx += 2

    # Validations dropdown pour versions
    col_idx = 2
    for table_name in consent_table_names:
        dv = DataValidation(type="list", formula1=f'INDIRECT("{table_name}[Libellé]")', allow_blank=True)
        dv.error = "Choisir une version du tableau Settings"
        dv.errorTitle = "Version invalide"
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{num_patients + 1}")
        col_idx += 2

    # Validation Note info
    note_col = 2 + len(consent_types) * 2
    _add_validation(ws, "list", '"Oui,Non"', f"{get_column_letter(note_col)}2:{get_column_letter(note_col)}{num_patients + 1}")

    # Tableau
    _create_table(ws, "Documents", f"A1:{get_column_letter(len(headers))}{num_patients + 1}", "TableStyleMedium3")

    # Mise en forme conditionnelle
    col_idx = 2
    for _ in consent_types:
        version_col = get_column_letter(col_idx)
        range_str = f"{version_col}2:{version_col}{num_patients + 1}"
        ws.conditional_formatting.add(range_str, FormulaRule(formula=[f'{version_col}2<>""'], fill=FILL_GREEN_PALE))
        col_idx += 2


def _create_adverse_events_sheet(ws: Worksheet) -> None:
    """Crée l'onglet Événements Indésirables."""
    headers = [
        "N° Patient", "N° EI", "Date survenue", "Date fin", "Description",
        "Gravité", "Sévérité", "Lien traitement", "Action prise",
        "Évolution", "EIG", "Date déclaration", "Commentaires"
    ]
    widths = [12, 8, 14, 14, 30, 12, 12, 14, 16, 12, 8, 14, 30]
    num_rows = 100

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell, FILL_RED)

    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Lignes et formats
    date_cols = [3, 4, 12]
    for row in range(2, num_rows + 2):
        ws.cell(row=row, column=1, value="")
        for col in date_cols:
            ws.cell(row=row, column=col).number_format = DATE_FORMAT_US

    # Validations
    _add_validation(ws, "list", '"Léger,Modéré,Sévère"', f"F2:F{num_rows + 1}")
    _add_validation(ws, "list", '"Grade 1,Grade 2,Grade 3,Grade 4,Grade 5"', f"G2:G{num_rows + 1}")
    _add_validation(ws, "list", '"Non lié,Peu probable,Possible,Probable,Certain"', f"H2:H{num_rows + 1}")
    _add_validation(ws, "list", '"Aucune,Traitement symptomatique,Arrêt temporaire,Arrêt définitif,Hospitalisation"', f"I2:I{num_rows + 1}")
    _add_validation(ws, "list", '"En cours,Résolu,Résolu avec séquelles,Décès"', f"J2:J{num_rows + 1}")
    _add_validation(ws, "list", '"Oui,Non"', f"K2:K{num_rows + 1}")

    _create_table(ws, "EvenementsIndesirables", f"A1:M{num_rows + 1}", "TableStyleMedium3")


def _create_treatment_sheet(ws: Worksheet) -> None:
    """Crée l'onglet Traitement IP."""
    headers = [
        "N° Patient", "Visite", "Date dispensation", "N° Lot",
        "Quantité dispensée", "Quantité retournée", "Compliance (%)", "Commentaires"
    ]
    widths = [12, 10, 16, 14, 18, 18, 14, 30]
    num_rows = 200

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell, FILL_PURPLE)

    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    for row in range(2, num_rows + 2):
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=3).number_format = DATE_FORMAT_US
        ws.cell(row=row, column=7).number_format = "0.0%"

    _create_table(ws, "TraitementIP", f"A1:H{num_rows + 1}", "TableStyleMedium5")


def _create_queries_sheet(ws: Worksheet) -> None:
    """Crée l'onglet Queries."""
    headers = [
        "N° Query", "N° Patient", "Visite", "Champ CRF", "Description",
        "Date ouverture", "Date réponse", "Date résolution", "Statut",
        "Réponse site", "Commentaires DM"
    ]
    widths = [10, 12, 10, 16, 30, 14, 14, 14, 12, 30, 30]
    num_rows = 200

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell, FILL_ORANGE)

    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    date_cols = [6, 7, 8]
    for row in range(2, num_rows + 2):
        ws.cell(row=row, column=1, value="")
        for col in date_cols:
            ws.cell(row=row, column=col).number_format = DATE_FORMAT_US

    _add_validation(ws, "list", '"Ouverte,En attente réponse,Résolue,Annulée"', f"I2:I{num_rows + 1}")

    _create_table(ws, "Queries", f"A1:K{num_rows + 1}", "TableStyleMedium3")

    # Mise en forme conditionnelle
    ws.conditional_formatting.add(f"I2:I{num_rows + 1}", FormulaRule(formula=['I2="Ouverte"'], fill=FILL_RED_LIGHT))
    ws.conditional_formatting.add(f"I2:I{num_rows + 1}", FormulaRule(formula=['I2="Résolue"'], fill=FILL_GREEN_LIGHT))


def _create_monitoring_sheet(ws: Worksheet, num_patients: int) -> None:
    """Crée l'onglet Monitoring."""
    headers = [
        "N° Patient", "Date dernière visite", "CRF complétés", "CRF en attente",
        "Taux complétion (%)", "Queries ouvertes", "Queries résolues",
        "Documents manquants", "Commentaires"
    ]
    widths = [12, 18, 14, 14, 16, 16, 16, 20, 30]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell, FILL_GREEN_DARK)

    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    for row in range(2, num_patients + 2):
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=2).number_format = DATE_FORMAT_US
        ws.cell(row=row, column=5).number_format = "0.0%"

    _create_table(ws, "Monitoring", f"A1:I{num_patients + 1}", "TableStyleMedium4")


def _create_dashboard_sheet(ws: Worksheet) -> None:
    """Crée l'onglet Dashboard avec compteurs automatiques."""
    # Section Recrutement
    ws["A1"] = "RECRUTEMENT"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = FILL_BLUE
    ws.merge_cells("A1:C1")

    metrics_recruitment = [
        ("Patients screenés", '=COUNTIF(\'Suivi Patients\'!H:H,"Screening")+COUNTIF(\'Suivi Patients\'!H:H,"Inclus")+COUNTIF(\'Suivi Patients\'!H:H,"Actif")+COUNTIF(\'Suivi Patients\'!H:H,"Terminé")+COUNTIF(\'Suivi Patients\'!H:H,"Sorti")'),
        ("Patients inclus", '=COUNTIF(\'Suivi Patients\'!H:H,"Inclus")+COUNTIF(\'Suivi Patients\'!H:H,"Actif")+COUNTIF(\'Suivi Patients\'!H:H,"Terminé")+COUNTIF(\'Suivi Patients\'!H:H,"Sorti")'),
        ("Patients actifs", '=COUNTIF(\'Suivi Patients\'!H:H,"Actif")'),
        ("Patients terminés", '=COUNTIF(\'Suivi Patients\'!H:H,"Terminé")'),
        ("Patients sortis prématurément", '=COUNTIF(\'Suivi Patients\'!H:H,"Sorti")'),
    ]
    for i, (label, formula) in enumerate(metrics_recruitment, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=formula)

    # Section Sécurité
    ws["A8"] = "SÉCURITÉ"
    ws["A8"].font = TITLE_FONT
    ws["A8"].fill = FILL_RED
    ws.merge_cells("A8:C8")

    metrics_safety = [
        ("Total EI", "=COUNTA('Événements Indésirables'!A:A)-1"),
        ("Total EIG", '=COUNTIF(\'Événements Indésirables\'!K:K,"Oui")'),
        ("EI en cours", '=COUNTIF(\'Événements Indésirables\'!J:J,"En cours")'),
    ]
    for i, (label, formula) in enumerate(metrics_safety, start=9):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=formula)

    # Section Data Management
    ws["A13"] = "DATA MANAGEMENT"
    ws["A13"].font = TITLE_FONT
    ws["A13"].fill = FILL_ORANGE
    ws.merge_cells("A13:C13")

    metrics_dm = [
        ("Queries ouvertes", '=COUNTIF(Queries!I:I,"Ouverte")'),
        ("Queries en attente", '=COUNTIF(Queries!I:I,"En attente réponse")'),
        ("Queries résolues", '=COUNTIF(Queries!I:I,"Résolue")'),
        ("Total queries", "=COUNTA(Queries!A:A)-1"),
    ]
    for i, (label, formula) in enumerate(metrics_dm, start=14):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=formula)

    # Section Sorties d'étude
    ws["A19"] = "RAISONS DE SORTIE"
    ws["A19"].font = TITLE_FONT
    ws["A19"].fill = FILL_PURPLE
    ws.merge_cells("A19:C19")

    raisons = [
        "Fin prévue", "Retrait consentement", "Perdu de vue",
        "Événement indésirable", "Décès", "Décision investigateur", "Autre"
    ]
    for i, raison in enumerate(raisons, start=20):
        ws.cell(row=i, column=1, value=raison)
        ws.cell(row=i, column=2, value=f"=COUNTIF('Suivi Patients'!J:J,\"{raison}\")")

    _set_column_widths(ws, {"A": 28, "B": 12, "C": 12})


# =============================================================================
# Fonction principale
# =============================================================================

def create_visit_tracking(
    num_visits: int = 10,
    num_patients: int = 50,
    output_path: Optional[str] = None
) -> Workbook:
    """
    Crée un fichier complet de suivi d'étude clinique.

    Args:
        num_visits: Nombre de visites à prévoir
        num_patients: Nombre de lignes patients
        output_path: Chemin de sauvegarde (optionnel)

    Returns:
        Workbook généré
    """
    wb = Workbook()

    # Onglet 1 : Settings
    ws_settings = wb.active
    ws_settings.title = "Settings"
    _create_settings_sheet(ws_settings, num_visits)

    # Onglet 2 : Suivi Patients
    ws_patients = wb.create_sheet(title="Suivi Patients")
    _create_patients_sheet(ws_patients, num_visits, num_patients)

    # Onglet 3 : Documents
    ws_docs = wb.create_sheet(title="Documents")
    _create_documents_sheet(ws_docs, num_patients)

    # Onglet 4 : Événements Indésirables
    ws_ei = wb.create_sheet(title="Événements Indésirables")
    _create_adverse_events_sheet(ws_ei)

    # Onglet 5 : Traitement IP
    ws_ip = wb.create_sheet(title="Traitement IP")
    _create_treatment_sheet(ws_ip)

    # Onglet 6 : Queries
    ws_queries = wb.create_sheet(title="Queries")
    _create_queries_sheet(ws_queries)

    # Onglet 7 : Monitoring
    ws_monitoring = wb.create_sheet(title="Monitoring")
    _create_monitoring_sheet(ws_monitoring, num_patients)

    # Onglet 8 : Dashboard
    ws_dashboard = wb.create_sheet(title="Dashboard")
    _create_dashboard_sheet(ws_dashboard)

    # Sauvegarde optionnelle
    if output_path:
        save_workbook(wb, output_path)

    return wb
