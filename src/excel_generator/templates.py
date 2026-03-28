"""
Système de templates pour la génération Excel.
"""

import re
import json
from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.styles import Font, Alignment
from openpyxl.utils.cell import range_boundaries

from .generator import create_workbook, save_workbook, write_data, auto_column_width, style_header
from .styles import (
    apply_border, merge_cells, apply_number_format,
    apply_style_preset
)


# =============================================================================
# Moteur de templates
# =============================================================================

class ExcelTemplate:
    """
    Classe de base pour les templates Excel.
    """

    def __init__(self, name: str, description: str = ""):
        self.name = name
        self.description = description
        self.wb: Optional[Workbook] = None
        self.variables: Dict[str, Any] = {}

    def set_variables(self, variables: Dict[str, Any]) -> "ExcelTemplate":
        """
        Définit les variables du template.

        Args:
            variables: Dictionnaire des variables

        Returns:
            self pour chaînage
        """
        self.variables = variables
        return self

    def _substitute(self, text: str) -> str:
        """
        Remplace les variables {{var}} dans le texte.
        """
        if not isinstance(text, str):
            return text

        def replace_var(match):
            var_name = match.group(1).strip()
            value = self.variables.get(var_name, match.group(0))
            # Formater les dates
            if isinstance(value, (datetime, date)):
                return value.strftime("%d/%m/%Y")
            return str(value)

        return re.sub(r"\{\{(\w+)\}\}", replace_var, text)

    def _process_cell_value(self, value: Any) -> Any:
        """
        Traite une valeur de cellule (substitution + conversion).
        """
        if isinstance(value, str):
            substituted = self._substitute(value)
            # Tenter de convertir en nombre si c'est numérique
            try:
                if "." in substituted:
                    return float(substituted)
                return int(substituted)
            except ValueError:
                return substituted
        return value

    def generate(self) -> Workbook:
        """
        Génère le workbook. À surcharger dans les sous-classes.
        """
        raise NotImplementedError("Les sous-classes doivent implémenter generate()")

    def save(self, filepath: str | Path) -> Path:
        """
        Génère et sauvegarde le fichier.

        Args:
            filepath: Chemin de destination

        Returns:
            Chemin absolu du fichier créé
        """
        if self.wb is None:
            self.generate()
        return save_workbook(self.wb, filepath)


# =============================================================================
# Template: Tableau simple
# =============================================================================

class TableTemplate(ExcelTemplate):
    """
    Template pour créer un tableau simple avec en-têtes et données.
    """

    def __init__(
        self,
        headers: List[str],
        data: List[List[Any]],
        title: Optional[str] = None,
        sheet_name: str = "Données"
    ):
        super().__init__("table", "Tableau simple avec en-têtes")
        self.headers = headers
        self.data = data
        self.title = title
        self.sheet_name = sheet_name

    def generate(self) -> Workbook:
        self.wb = create_workbook(self.sheet_name)
        ws = self.wb.active

        start_row = 1

        # Titre optionnel
        if self.title:
            title_text = self._substitute(self.title)
            merge_cells(ws, 1, 1, 1, len(self.headers), title_text, bold=True, bg_color="4472C4")
            ws.row_dimensions[1].height = 30
            start_row = 3

        # En-têtes
        header_row = start_row
        for col, header in enumerate(self.headers, start=1):
            ws.cell(row=header_row, column=col, value=self._substitute(header))
        apply_style_preset(ws, header_row, 1, header_row, len(self.headers), "header")

        # Données
        for row_idx, row_data in enumerate(self.data, start=header_row + 1):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=self._process_cell_value(value))

        # Bordures et largeur auto
        end_row = header_row + len(self.data)
        apply_border(ws, header_row, 1, end_row, len(self.headers))
        auto_column_width(ws)

        return self.wb


# =============================================================================
# Template: Facture
# =============================================================================

class InvoiceTemplate(ExcelTemplate):
    """
    Template de facture.

    Variables attendues:
        - numero: Numéro de facture
        - date: Date de facture
        - client_nom: Nom du client
        - client_adresse: Adresse du client
        - entreprise_nom: Nom de l'entreprise
        - entreprise_adresse: Adresse de l'entreprise
        - lignes: Liste de dict avec {description, quantite, prix_unitaire}
        - tva: Taux de TVA (ex: 0.20 pour 20%)
    """

    def __init__(self):
        super().__init__("facture", "Facture professionnelle")

    def generate(self) -> Workbook:
        self.wb = create_workbook("Facture")
        ws = self.wb.active

        # Configuration page
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

        # En-tête entreprise
        row = 1
        entreprise = self.variables.get("entreprise_nom", "Entreprise")
        ws.cell(row=row, column=1, value=entreprise).font = Font(bold=True, size=16)
        row += 1
        ws.cell(row=row, column=1, value=self.variables.get("entreprise_adresse", ""))
        row += 2

        # Titre FACTURE
        ws.cell(row=row, column=1, value="FACTURE").font = Font(bold=True, size=14)
        ws.cell(row=row, column=3, value=f"N° {self.variables.get('numero', 'XXX')}")
        row += 1
        date_facture = self.variables.get("date", datetime.now())
        if isinstance(date_facture, str):
            ws.cell(row=row, column=3, value=f"Date: {date_facture}")
        else:
            ws.cell(row=row, column=3, value=f"Date: {date_facture.strftime('%d/%m/%Y')}")
        row += 2

        # Client
        ws.cell(row=row, column=1, value="Client:").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value=self.variables.get("client_nom", ""))
        row += 1
        ws.cell(row=row, column=1, value=self.variables.get("client_adresse", ""))
        row += 2

        # Tableau des lignes
        header_row = row
        headers = ["Description", "Quantité", "Prix unitaire", "Total"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=h)
        apply_style_preset(ws, row, 1, row, 4, "header")
        row += 1

        # Lignes de facture
        lignes = self.variables.get("lignes", [])
        sous_total = 0

        for ligne in lignes:
            desc = ligne.get("description", "")
            qte = ligne.get("quantite", 1)
            prix = ligne.get("prix_unitaire", 0)
            total_ligne = qte * prix
            sous_total += total_ligne

            ws.cell(row=row, column=1, value=desc)
            ws.cell(row=row, column=2, value=qte)
            ws.cell(row=row, column=3, value=prix)
            ws.cell(row=row, column=4, value=total_ligne)
            row += 1

        # Bordures tableau
        apply_border(ws, header_row, 1, row - 1, 4)

        # Formats monétaires
        apply_number_format(ws, header_row + 1, 3, row - 1, 4, "euro")

        row += 1

        # Totaux
        tva_taux = self.variables.get("tva", 0.20)
        tva_montant = sous_total * tva_taux
        total_ttc = sous_total + tva_montant

        ws.cell(row=row, column=3, value="Sous-total HT:")
        ws.cell(row=row, column=4, value=sous_total)
        row += 1

        ws.cell(row=row, column=3, value=f"TVA ({int(tva_taux * 100)}%):")
        ws.cell(row=row, column=4, value=tva_montant)
        row += 1

        ws.cell(row=row, column=3, value="TOTAL TTC:").font = Font(bold=True)
        cell_total = ws.cell(row=row, column=4, value=total_ttc)
        cell_total.font = Font(bold=True)

        apply_number_format(ws, row - 2, 4, row, 4, "euro")

        return self.wb




# =============================================================================
# Template: Budget mensuel
# =============================================================================

class BudgetTemplate(ExcelTemplate):
    """
    Template de budget mensuel.

    Variables attendues:
        - mois: Mois du budget (ex: "Mars 2026")
        - revenus: Liste de dict {categorie, montant}
        - depenses: Liste de dict {categorie, montant}
    """

    def __init__(self):
        super().__init__("budget", "Budget mensuel")

    def generate(self) -> Workbook:
        self.wb = create_workbook("Budget")
        ws = self.wb.active

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 5
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 15

        # Titre
        mois = self.variables.get("mois", datetime.now().strftime("%B %Y"))
        merge_cells(ws, 1, 1, 1, 5, f"Budget - {mois}", bold=True, bg_color="4472C4")
        ws.row_dimensions[1].height = 30

        row = 3

        # Revenus
        ws.cell(row=row, column=1, value="REVENUS").font = Font(bold=True, color="006100")
        apply_style_preset(ws, row, 1, row, 2, "success")
        row += 1

        revenus = self.variables.get("revenus", [])
        total_revenus = 0
        for rev in revenus:
            ws.cell(row=row, column=1, value=rev.get("categorie", ""))
            montant = rev.get("montant", 0)
            ws.cell(row=row, column=2, value=montant)
            total_revenus += montant
            row += 1

        ws.cell(row=row, column=1, value="Total Revenus").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_revenus).font = Font(bold=True)
        apply_style_preset(ws, row, 1, row, 2, "total")
        row += 2

        # Dépenses
        ws.cell(row=row, column=1, value="DÉPENSES").font = Font(bold=True, color="9C0006")
        apply_style_preset(ws, row, 1, row, 2, "attention")
        row += 1

        depenses = self.variables.get("depenses", [])
        total_depenses = 0
        for dep in depenses:
            ws.cell(row=row, column=1, value=dep.get("categorie", ""))
            montant = dep.get("montant", 0)
            ws.cell(row=row, column=2, value=montant)
            total_depenses += montant
            row += 1

        ws.cell(row=row, column=1, value="Total Dépenses").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_depenses).font = Font(bold=True)
        apply_style_preset(ws, row, 1, row, 2, "total")
        row += 2

        # Solde
        solde = total_revenus - total_depenses
        ws.cell(row=row, column=1, value="SOLDE").font = Font(bold=True, size=12)
        cell_solde = ws.cell(row=row, column=2, value=solde)
        cell_solde.font = Font(bold=True, size=12)

        if solde >= 0:
            apply_style_preset(ws, row, 1, row, 2, "success")
        else:
            apply_style_preset(ws, row, 1, row, 2, "attention")

        # Format monétaire
        apply_number_format(ws, 4, 2, row, 2, "euro")

        return self.wb


# =============================================================================
# Template: Planning hebdomadaire
# =============================================================================

class PlanningTemplate(ExcelTemplate):
    """
    Template de planning hebdomadaire.

    Variables attendues:
        - semaine: Numéro ou description de la semaine
        - horaires: Liste des créneaux horaires (ex: ["8h", "9h", ...])
        - taches: Dict {jour: {heure: tache}} (optionnel)
    """

    def __init__(self):
        super().__init__("planning", "Planning hebdomadaire")

    def generate(self) -> Workbook:
        self.wb = create_workbook("Planning")
        ws = self.wb.active

        jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
        horaires = self.variables.get("horaires", [f"{h}h" for h in range(8, 19)])
        taches = self.variables.get("taches", {})

        # Largeurs
        ws.column_dimensions["A"].width = 8
        for i in range(len(jours)):
            ws.column_dimensions[chr(66 + i)].width = 15

        # Titre
        semaine = self.variables.get("semaine", "")
        titre = f"Planning{' - ' + str(semaine) if semaine else ''}"
        merge_cells(ws, 1, 1, 1, len(jours) + 1, titre, bold=True, bg_color="4472C4")

        # En-têtes jours
        row = 2
        ws.cell(row=row, column=1, value="Heure")
        for col, jour in enumerate(jours, start=2):
            ws.cell(row=row, column=col, value=jour)
        apply_style_preset(ws, row, 1, row, len(jours) + 1, "header")

        # Grille horaires
        for row_idx, heure in enumerate(horaires, start=3):
            ws.cell(row=row_idx, column=1, value=heure)
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")

            # Remplir les tâches si définies
            for col_idx, jour in enumerate(jours, start=2):
                if jour in taches and heure in taches[jour]:
                    ws.cell(row=row_idx, column=col_idx, value=taches[jour][heure])

        # Bordures
        end_row = 2 + len(horaires)
        apply_border(ws, 2, 1, end_row, len(jours) + 1)

        # Hauteur des lignes
        for r in range(3, end_row + 1):
            ws.row_dimensions[r].height = 25

        return self.wb


# =============================================================================
# Registre des templates
# =============================================================================

TEMPLATES = {
    "facture": InvoiceTemplate,
    "budget": BudgetTemplate,
    "planning": PlanningTemplate,
}


def get_template(name: str) -> ExcelTemplate:
    """
    Récupère une instance de template par son nom.

    Args:
        name: Nom du template

    Returns:
        Instance du template

    Raises:
        ValueError: Si le template n'existe pas
    """
    if name not in TEMPLATES:
        available = ", ".join(TEMPLATES.keys())
        raise ValueError(f"Template '{name}' inconnu. Disponibles: {available}")

    return TEMPLATES[name]()


def list_templates() -> List[Dict[str, str]]:
    """
    Liste tous les templates disponibles.

    Returns:
        Liste de dict avec name et description
    """
    result = []
    for name, cls in TEMPLATES.items():
        instance = cls()
        result.append({
            "name": name,
            "description": instance.description
        })
    return result


def create_from_dict(
    data: Dict[str, Any],
    title: Optional[str] = None,
    sheet_name: str = "Données"
) -> Workbook:
    """
    Crée un fichier Excel à partir d'un dictionnaire.

    Args:
        data: Dictionnaire avec les clés comme en-têtes
        title: Titre optionnel
        sheet_name: Nom de la feuille

    Returns:
        Workbook généré
    """
    if not data:
        raise ValueError("Le dictionnaire est vide")

    # Convertir dict en headers + data
    headers = list(data.keys())

    # Déterminer le nombre de lignes
    first_value = list(data.values())[0]
    if isinstance(first_value, list):
        num_rows = len(first_value)
        rows = []
        for i in range(num_rows):
            row = [data[h][i] if isinstance(data[h], list) and i < len(data[h]) else data[h] for h in headers]
            rows.append(row)
    else:
        rows = [[data[h] for h in headers]]

    template = TableTemplate(headers, rows, title, sheet_name)
    return template.generate()


def create_from_json(
    json_str: str,
    title: Optional[str] = None,
    sheet_name: str = "Données"
) -> Workbook:
    """
    Crée un fichier Excel à partir d'une chaîne JSON.

    Args:
        json_str: JSON (dict ou liste de dicts)
        title: Titre optionnel
        sheet_name: Nom de la feuille

    Returns:
        Workbook généré
    """
    data = json.loads(json_str)

    if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
        # Liste de dicts -> tableau
        headers = list(data[0].keys())
        rows = [[item.get(h, "") for h in headers] for item in data]
        template = TableTemplate(headers, rows, title, sheet_name)
        return template.generate()
    elif isinstance(data, dict):
        return create_from_dict(data, title, sheet_name)
    else:
        raise ValueError("Format JSON non supporté. Attendu: dict ou liste de dicts")
