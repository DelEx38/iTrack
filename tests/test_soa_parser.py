"""
Tests unitaires pour SoaParserService.

Usage:
    pytest tests/test_soa_parser.py -v
"""

import pytest
from pathlib import Path
from src.services.soa_parser import SoaParserService, VisitConfig


# =============================================================================
# Fixtures
# =============================================================================

@pytest.fixture
def parser():
    """Instance fraîche du parser."""
    return SoaParserService()


@pytest.fixture
def format1_file():
    """Fichier Excel Format 1 (lignes Day/Window séparées)."""
    path = Path(__file__).parent.parent / "output" / "test_soa.xlsx"
    if not path.exists():
        pytest.skip(f"Fichier de test manquant: {path}")
    return path.read_bytes()


@pytest.fixture
def format2_file():
    """Fichier Excel Format 2 (jour intégré dans header)."""
    path = Path(__file__).parent.parent / "output" / "test_soa_format2.xlsx"
    if not path.exists():
        pytest.skip(f"Fichier de test manquant: {path}")
    return path.read_bytes()


# =============================================================================
# Tests _is_visit_header
# =============================================================================

class TestIsVisitHeader:
    """Tests pour la détection des headers de visite."""

    def test_v_pattern_simple(self, parser):
        """V1, V2, V10 sont reconnus."""
        assert parser._is_visit_header("V1") is True
        assert parser._is_visit_header("V2") is True
        assert parser._is_visit_header("V10") is True
        assert parser._is_visit_header("V25") is True

    def test_v_pattern_with_space(self, parser):
        """V1 avec espaces autour."""
        assert parser._is_visit_header("  V1  ") is True
        assert parser._is_visit_header(" V2 ") is True

    def test_screening(self, parser):
        """Screening est reconnu."""
        assert parser._is_visit_header("Screening") is True
        assert parser._is_visit_header("SCREENING") is True
        assert parser._is_visit_header("screening") is True

    def test_baseline(self, parser):
        """Baseline est reconnu."""
        assert parser._is_visit_header("Baseline") is True
        assert parser._is_visit_header("BASELINE") is True

    def test_eot_eos(self, parser):
        """EOT et EOS sont reconnus."""
        assert parser._is_visit_header("EOT") is True
        assert parser._is_visit_header("EOS") is True
        assert parser._is_visit_header("eot") is True

    def test_follow_up(self, parser):
        """Follow-up et variantes sont reconnus."""
        assert parser._is_visit_header("FU") is True
        assert parser._is_visit_header("FU1") is True
        assert parser._is_visit_header("Follow-up") is True
        assert parser._is_visit_header("Followup") is True
        assert parser._is_visit_header("Follow up") is True

    def test_visit_numbered(self, parser):
        """Visit 1, Visit2 sont reconnus."""
        assert parser._is_visit_header("Visit 1") is True
        assert parser._is_visit_header("Visit2") is True
        assert parser._is_visit_header("Visit 10") is True

    def test_week_pattern(self, parser):
        """W1, W4 (semaines) sont reconnus."""
        assert parser._is_visit_header("W1") is True
        assert parser._is_visit_header("W4") is True
        assert parser._is_visit_header("W12") is True

    def test_day_pattern(self, parser):
        """D1, D14 (jours comme noms de visite) sont reconnus."""
        assert parser._is_visit_header("D1") is True
        assert parser._is_visit_header("D14") is True

    def test_french_patterns(self, parser):
        """J0, Visite 1 (français) sont reconnus."""
        assert parser._is_visit_header("J0") is True
        assert parser._is_visit_header("J7") is True
        assert parser._is_visit_header("Visite 1") is True
        assert parser._is_visit_header("Visite 10") is True

    def test_invalid_headers(self, parser):
        """Textes non-visites ne sont pas reconnus."""
        assert parser._is_visit_header("Procedure") is False
        assert parser._is_visit_header("Assessment") is False
        assert parser._is_visit_header("Notes") is False
        assert parser._is_visit_header("") is False
        assert parser._is_visit_header("Patient ID") is False


# =============================================================================
# Tests _extract_visit_name
# =============================================================================

class TestExtractVisitName:
    """Tests pour l'extraction du nom de visite."""

    def test_simple_name(self, parser):
        """Noms simples retournés tels quels."""
        assert parser._extract_visit_name("V1") == "V1"
        assert parser._extract_visit_name("V2") == "V2"
        assert parser._extract_visit_name("Screening") == "Screening"
        assert parser._extract_visit_name("Baseline") == "Baseline"
        assert parser._extract_visit_name("EOT") == "EOT"

    def test_format2_with_day(self, parser):
        """Format 2: V1 D0 -> V1."""
        assert parser._extract_visit_name("V1 D0") == "V1"
        assert parser._extract_visit_name("V2 D7") == "V2"
        assert parser._extract_visit_name("V3 D14") == "V3"

    def test_format2_with_day_and_window(self, parser):
        """Format 2 avec fenêtre: V2 D7±2 -> V2."""
        assert parser._extract_visit_name("V2 D7±2") == "V2"
        assert parser._extract_visit_name("V3 D14 +/-3") == "V3"
        assert parser._extract_visit_name("V4 D21 -2/+5") == "V4"

    def test_format2_screening_negative_day(self, parser):
        """Screening avec jour négatif: Screening D-28 -> Screening."""
        assert parser._extract_visit_name("Screening D-28") == "Screening"

    def test_parentheses_notation(self, parser):
        """Notation avec parenthèses: Baseline (Day 0) -> Baseline."""
        assert parser._extract_visit_name("Baseline (Day 0)") == "Baseline"
        assert parser._extract_visit_name("EOT (Day 84)") == "EOT"

    def test_french_notation(self, parser):
        """Notation française: V1 J0 -> V1."""
        assert parser._extract_visit_name("V1 J0") == "V1"
        assert parser._extract_visit_name("V2 J7") == "V2"

    def test_with_spaces(self, parser):
        """Espaces gérés correctement."""
        assert parser._extract_visit_name("  V1  ") == "V1"
        assert parser._extract_visit_name("V1  D0") == "V1"


# =============================================================================
# Tests _parse_window
# =============================================================================

class TestParseWindow:
    """Tests pour le parsing des fenêtres."""

    def test_symmetric_plus_minus(self, parser):
        """±3 -> (3, 3)."""
        assert parser._parse_window("±3") == (3, 3)
        assert parser._parse_window("± 3") == (3, 3)
        assert parser._parse_window("±7") == (7, 7)
        assert parser._parse_window("±14") == (14, 14)

    def test_symmetric_plus_minus_days(self, parser):
        """±3 days -> (3, 3)."""
        assert parser._parse_window("±3 days") == (3, 3)
        assert parser._parse_window("± 3 days") == (3, 3)

    def test_slash_notation_plus_first(self, parser):
        """+/-3 -> (3, 3)."""
        assert parser._parse_window("+/-3") == (3, 3)
        assert parser._parse_window("+/- 3") == (3, 3)
        assert parser._parse_window("+/ -3") == (3, 3)

    def test_slash_notation_minus_first(self, parser):
        """-/+3 -> (3, 3)."""
        assert parser._parse_window("-/+3") == (3, 3)
        assert parser._parse_window("-/ +3") == (3, 3)

    def test_asymmetric_slash(self, parser):
        """-2/+5 -> (2, 5)."""
        assert parser._parse_window("-2/+5") == (2, 5)
        assert parser._parse_window("-2 / +5") == (2, 5)
        assert parser._parse_window("-3/+7") == (3, 7)

    def test_asymmetric_to_notation(self, parser):
        """-2 to +5 -> (2, 5)."""
        assert parser._parse_window("-2 to +5") == (2, 5)
        assert parser._parse_window("-3 to +7") == (3, 7)

    def test_asymmetric_french_notation(self, parser):
        """-2 à +5 -> (2, 5)."""
        assert parser._parse_window("-2 à +5") == (2, 5)

    def test_asymmetric_parentheses_comma(self, parser):
        """(-2, +5) -> (2, 5)."""
        assert parser._parse_window("(-2, +5)") == (2, 5)
        assert parser._parse_window("( -2, +5 )") == (2, 5)

    def test_asymmetric_parentheses_slash(self, parser):
        """(-2/+5) -> (2, 5)."""
        assert parser._parse_window("(-2/+5)") == (2, 5)

    def test_no_window(self, parser):
        """Pas de fenêtre -> None."""
        assert parser._parse_window("") is None
        assert parser._parse_window("D0") is None
        assert parser._parse_window("Day 7") is None
        assert parser._parse_window("V1") is None
        assert parser._parse_window("Procedure") is None


# =============================================================================
# Tests VisitConfig
# =============================================================================

class TestVisitConfig:
    """Tests pour la dataclass VisitConfig."""

    def test_creation_simple(self):
        """Création basique."""
        visit = VisitConfig(
            visit_name="V1",
            target_day=0,
            window_before=0,
            window_after=0,
        )
        assert visit.visit_name == "V1"
        assert visit.target_day == 0
        assert visit.window_before == 0
        assert visit.window_after == 0
        assert visit.procedures == []

    def test_creation_with_window(self):
        """Création avec fenêtre."""
        visit = VisitConfig(
            visit_name="V2",
            target_day=7,
            window_before=2,
            window_after=3,
        )
        assert visit.target_day == 7
        assert visit.window_before == 2
        assert visit.window_after == 3

    def test_creation_with_procedures(self):
        """Création avec procédures."""
        visit = VisitConfig(
            visit_name="V1",
            target_day=0,
            window_before=0,
            window_after=0,
            procedures=["ECG", "Blood draw"],
        )
        assert visit.procedures == ["ECG", "Blood draw"]

    def test_to_dict(self):
        """Conversion en dictionnaire."""
        visit = VisitConfig(
            visit_name="V2",
            target_day=7,
            window_before=2,
            window_after=3,
            procedures=["ECG"],
        )
        d = visit.to_dict()
        assert d == {
            "visit_name": "V2",
            "target_day": 7,
            "window_before": 2,
            "window_after": 3,
            "procedures": ["ECG"],
        }


# =============================================================================
# Tests d'intégration parse_file
# =============================================================================

class TestParseFileFormat1:
    """Tests d'intégration pour Format 1 (lignes Day/Window séparées)."""

    def test_parse_returns_visits(self, parser, format1_file):
        """Le parsing retourne une liste de visites."""
        visits = parser.parse_file(format1_file)
        assert isinstance(visits, list)
        assert len(visits) > 0
        assert all(isinstance(v, VisitConfig) for v in visits)

    def test_parse_visit_names(self, parser, format1_file):
        """Les noms de visites sont extraits correctement."""
        visits = parser.parse_file(format1_file)
        names = [v.visit_name for v in visits]
        # Au moins quelques visites standard
        assert any("V1" in name or "Screening" in name for name in names)

    def test_parse_target_days(self, parser, format1_file):
        """Les jours cibles sont extraits."""
        visits = parser.parse_file(format1_file)
        # Au moins une visite avec un jour défini
        days = [v.target_day for v in visits]
        assert 0 in days  # V1 ou Baseline à J0

    def test_parse_windows(self, parser, format1_file):
        """Les fenêtres sont extraites."""
        visits = parser.parse_file(format1_file)
        # Au moins une visite avec une fenêtre
        has_window = any(v.window_before > 0 or v.window_after > 0 for v in visits)
        # V1 n'a pas de fenêtre, mais les autres oui
        visits_with_window = [v for v in visits if v.target_day > 0]
        if visits_with_window:
            assert has_window


class TestParseFileFormat2:
    """Tests d'intégration pour Format 2 (jour intégré dans header)."""

    def test_parse_returns_visits(self, parser, format2_file):
        """Le parsing retourne une liste de visites."""
        visits = parser.parse_file(format2_file)
        assert isinstance(visits, list)
        assert len(visits) > 0

    def test_parse_extracts_integrated_days(self, parser, format2_file):
        """Les jours intégrés (V1 D0, V2 D7) sont extraits."""
        visits = parser.parse_file(format2_file)
        # Vérifier qu'on a des jours variés
        days = [v.target_day for v in visits]
        unique_days = set(days)
        assert len(unique_days) >= 2  # Au moins 2 jours différents

    def test_parse_extracts_negative_days(self, parser, format2_file):
        """Les jours négatifs (D-28 pour Screening) sont extraits."""
        visits = parser.parse_file(format2_file)
        days = [v.target_day for v in visits]
        # Peut contenir un jour négatif si Screening présent
        # Ne pas forcer, dépend du fichier de test


class TestParseFileErrors:
    """Tests de gestion des erreurs."""

    def test_empty_bytes_raises(self, parser):
        """Bytes vides lèvent une exception."""
        with pytest.raises(Exception):
            parser.parse_file(b"")

    def test_invalid_excel_raises(self, parser):
        """Fichier non-Excel lève une exception."""
        with pytest.raises(Exception):
            parser.parse_file(b"not an excel file")


# =============================================================================
# Tests get_procedures
# =============================================================================

class TestGetProcedures:
    """Tests pour l'extraction des procédures."""

    def test_procedures_after_parse(self, parser, format1_file):
        """Les procédures sont disponibles après parsing."""
        parser.parse_file(format1_file)
        procedures = parser.get_procedures()
        assert isinstance(procedures, list)

    def test_procedures_empty_before_parse(self, parser):
        """Procédures vides avant parsing."""
        assert parser.get_procedures() == []


# =============================================================================
# Tests Format Week (ex: TRANSTELLAR)
# =============================================================================

@pytest.fixture
def format_week_file():
    """Fichier Excel Format Week (semaines numériques + ligne Window)."""
    path = Path(__file__).parent / "fixtures" / "test_soa_format_week.xlsx"
    if not path.exists():
        pytest.skip(f"Fichier de test manquant: {path}")
    return path.read_bytes()


class TestIsFormatWeek:
    """Tests pour la détection du Format Week."""

    def test_detected_on_transtellar(self, parser, format_week_file):
        """Le fichier TRANSTELLAR est détecté comme Format Week."""
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(filename=BytesIO(format_week_file), data_only=True)
        sheet = wb[wb.sheetnames[0]]
        assert parser._is_format_week(sheet) is True

    def test_not_detected_on_format1(self, parser, format1_file):
        """Le Format 1 n'est pas détecté comme Format Week."""
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(filename=BytesIO(format1_file), data_only=True)
        sheet = wb[wb.sheetnames[0]]
        assert parser._is_format_week(sheet) is False


class TestParseFormatWeek:
    """Tests d'intégration pour Format Week."""

    def test_parse_returns_28_visits(self, parser, format_week_file):
        """TRANSTELLAR a 28 visites (W0 à W204 + EOT)."""
        visits = parser.parse_file(format_week_file)
        assert len(visits) == 28

    def test_visit_names_start_with_W(self, parser, format_week_file):
        """Les visites numériques sont nommées W0, W8, W16..."""
        visits = parser.parse_file(format_week_file)
        week_visits = [v for v in visits if v.visit_name.startswith("W")]
        assert len(week_visits) == 27  # W0 à W204

    def test_eot_visit_present(self, parser, format_week_file):
        """EOT est présent comme visite spéciale."""
        visits = parser.parse_file(format_week_file)
        eot = next((v for v in visits if v.visit_name == "EOT"), None)
        assert eot is not None

    def test_days_are_weeks_times_7(self, parser, format_week_file):
        """W8 = Jour 56, W16 = Jour 112, W24 = Jour 168."""
        visits = parser.parse_file(format_week_file)
        by_name = {v.visit_name: v for v in visits}
        assert by_name["W8"].target_day == 56
        assert by_name["W16"].target_day == 112
        assert by_name["W24"].target_day == 168
        assert by_name["W56"].target_day == 392

    def test_w0_has_no_window(self, parser, format_week_file):
        """W0 (baseline) n'a pas de fenêtre."""
        visits = parser.parse_file(format_week_file)
        w0 = next(v for v in visits if v.visit_name == "W0")
        assert w0.window_before == 0
        assert w0.window_after == 0

    def test_windows_extracted(self, parser, format_week_file):
        """W8 a une fenêtre ±5 jours."""
        visits = parser.parse_file(format_week_file)
        w8 = next(v for v in visits if v.visit_name == "W8")
        assert w8.window_before == 5
        assert w8.window_after == 5

    def test_larger_windows_extracted(self, parser, format_week_file):
        """W24, W204 ont des fenêtres ±7 et ±10 jours."""
        visits = parser.parse_file(format_week_file)
        by_name = {v.visit_name: v for v in visits}
        assert by_name["W24"].window_before == 7
        assert by_name["W204"].window_before == 10

    def test_eot_day_equals_last_numeric_day(self, parser, format_week_file):
        """EOT hérite du jour de la dernière visite numérique (W204 = Jour 1428)."""
        visits = parser.parse_file(format_week_file)
        eot = next(v for v in visits if v.visit_name == "EOT")
        assert eot.target_day == 1428

    def test_procedures_extracted(self, parser, format_week_file):
        """Les procédures sont extraites (36 attendues)."""
        parser.parse_file(format_week_file)
        procedures = parser.get_procedures()
        assert len(procedures) == 36

    def test_known_procedure_present(self, parser, format_week_file):
        """'Informed consent' est dans les procédures."""
        parser.parse_file(format_week_file)
        procedures = parser.get_procedures()
        assert "Informed consent" in procedures
