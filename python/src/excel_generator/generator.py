"""
Fonctions de génération de fichiers Excel.
"""

from pathlib import Path
from typing import List, Any, Optional, Union
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_workbook(sheet_name: str = "Feuille1") -> Workbook:
    """
    Crée un nouveau classeur Excel.

    Args:
        sheet_name: Nom de la première feuille

    Returns:
        Workbook openpyxl
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    return wb


def save_workbook(wb: Workbook, filepath: Union[str, Path]) -> Path:
    """
    Sauvegarde le classeur Excel.

    Args:
        wb: Classeur à sauvegarder
        filepath: Chemin du fichier (avec ou sans extension .xlsx)

    Returns:
        Chemin absolu du fichier sauvegardé
    """
    filepath = Path(filepath)
    if filepath.suffix.lower() != ".xlsx":
        filepath = filepath.with_suffix(".xlsx")

    # Créer le dossier parent si nécessaire
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb.save(filepath)
    return filepath.resolve()


def add_sheet(wb: Workbook, name: str) -> Worksheet:
    """
    Ajoute une nouvelle feuille au classeur.

    Args:
        wb: Classeur
        name: Nom de la feuille

    Returns:
        La nouvelle feuille créée
    """
    return wb.create_sheet(title=name)


def write_row(ws: Worksheet, row: int, data: List[Any]) -> None:
    """
    Écrit une ligne de données.

    Args:
        ws: Feuille de calcul
        row: Numéro de ligne (1-indexed)
        data: Liste des valeurs à écrire
    """
    for col, value in enumerate(data, start=1):
        ws.cell(row=row, column=col, value=value)


def write_data(
    ws: Worksheet,
    data: List[List[Any]],
    start_row: int = 1,
    start_col: int = 1
) -> None:
    """
    Écrit un bloc de données (plusieurs lignes).

    Args:
        ws: Feuille de calcul
        data: Liste de lignes (chaque ligne est une liste de valeurs)
        start_row: Ligne de départ (1-indexed)
        start_col: Colonne de départ (1-indexed)
    """
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            ws.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


def style_header(
    ws: Worksheet,
    row: int = 1,
    num_cols: Optional[int] = None,
    bold: bool = True,
    bg_color: str = "4472C4",
    font_color: str = "FFFFFF"
) -> None:
    """
    Applique un style d'en-tête à une ligne.

    Args:
        ws: Feuille de calcul
        row: Numéro de ligne
        num_cols: Nombre de colonnes à styliser (auto-détecté si None)
        bold: Texte en gras
        bg_color: Couleur de fond (hex sans #)
        font_color: Couleur du texte (hex sans #)
    """
    if num_cols is None:
        num_cols = ws.max_column

    header_font = Font(bold=bold, color=font_color)
    header_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align


def auto_column_width(ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
    """
    Ajuste automatiquement la largeur des colonnes.

    Args:
        ws: Feuille de calcul
        min_width: Largeur minimale
        max_width: Largeur maximale
    """
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width
